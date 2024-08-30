import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import webbrowser

# Function to create a new workbook or load an existing one
def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["Date Applied", "Company", "Position", "Status", "Job Link", "Job Site"])
    return wb, ws

# Function to add job application data
def add_job_application(filename, company, position, job_link, status, job_site):
    wb, ws = get_workbook(filename)
    date_applied = datetime.now().strftime("%d/%m/%Y")  # Change date format here
    ws.append([date_applied, company, position, status, job_link, job_site])
    wb.save(filename)
    messagebox.showinfo("Success", "Application logged successfully!")
    update_treeview()

# Function to update the treeview with entries
def update_treeview():
    wb, ws = get_workbook(filename)
    for row in tree.get_children():
        tree.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_applied = row[0]
        company = row[1]
        position = row[2]
        status = row[3]
        tree.insert("", tk.END, values=(date_applied, company, position, status))

# Function to handle the form submission
def submit():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()

    if company and position and job_link and status and job_site:
        add_job_application(filename, company, position, job_link, status, job_site)
        entry_company.delete(0, tk.END)
        entry_position.delete(0, tk.END)
        entry_job_link.delete(0, tk.END)
        entry_status.delete(0, tk.END)
        entry_job_site.delete(0, tk.END)
    else:
        messagebox.showwarning("Input Error", "All fields are required.")

# Function to edit selected entry
def edit_entry():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to edit.")
        return

    item = tree.item(selected_item)
    date_applied, company, position, status = item['values']

    field = field_var.get()
    wb, ws = get_workbook(filename)
    index = tree.index(selected_item) + 2  # offset for header row and 0-index

    if field == "Company":
        new_value = simpledialog.askstring("Edit Company", "Enter new company:", initialvalue=company)
        if new_value:
            tree.item(selected_item, values=(date_applied, new_value, position, status))
            ws.cell(row=index, column=2).value = new_value
    elif field == "Position":
        new_value = simpledialog.askstring("Edit Position", "Enter new position:", initialvalue=position)
        if new_value:
            tree.item(selected_item, values=(date_applied, company, new_value, status))
            ws.cell(row=index, column=3).value = new_value
    elif field == "Status":
        new_value = simpledialog.askstring("Edit Status", "Enter new status:", initialvalue=status)
        if new_value:
            tree.item(selected_item, values=(date_applied, company, position, new_value))
            ws.cell(row=index, column=4).value = new_value
    elif field == "Date":
        new_value = simpledialog.askstring("Edit Date", "Enter new date (DD/MM/YYYY):", initialvalue=date_applied)
        if new_value:
            tree.item(selected_item, values=(new_value, company, position, status))
            ws.cell(row=index, column=1).value = new_value
    elif field == "Job Link":
        new_value = simpledialog.askstring("Edit Job Link", "Enter new job link:", initialvalue=get_cell_value(index, 5))
        if new_value:
            ws.cell(row=index, column=5).value = new_value
    elif field == "Job Site":
        new_value = simpledialog.askstring("Edit Job Site", "Enter new job site:", initialvalue=get_cell_value(index, 6))
        if new_value:
            ws.cell(row=index, column=6).value = new_value
    else:
        messagebox.showwarning("Field Error", "Invalid field selected.")
        return

    wb.save(filename)
    messagebox.showinfo("Success", "Entry updated successfully!")

# Helper function to get cell value
def get_cell_value(row, col):
    wb, ws = get_workbook(filename)
    return ws.cell(row=row, column=col).value

# Function to delete selected entry
def delete_entry():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to delete.")
        return

    confirm = messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this entry?")
    if not confirm:
        return

    item = tree.item(selected_item)
    date_applied, company, position, status = item['values']

    wb, ws = get_workbook(filename)
    for row in ws.iter_rows(min_row=2):
        if (row[0].value == date_applied and
            row[1].value == company and
            row[2].value == position and
            row[3].value == status):
            ws.delete_rows(row[0].row)
            break
    wb.save(filename)
    update_treeview()
    messagebox.showinfo("Success", "Entry deleted successfully!")

# Function to open the link of the selected entry
def open_link():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to open.")
        return

    item = tree.item(selected_item)
    date_applied, company, position, status = item['values']
    
    wb, ws = get_workbook(filename)
    job_link = get_cell_value(tree.index(selected_item) + 2, 5)
    if job_link:
        webbrowser.open(job_link)
    else:
        messagebox.showwarning("Link Error", "No link available for the selected entry.")

# Function to toggle always on top
def toggle_always_on_top():
    global always_on_top
    always_on_top = not always_on_top
    root.attributes("-topmost", always_on_top)
    toggle_button.config(relief=tk.SUNKEN if always_on_top else tk.RAISED)

# GUI setup
filename = "job_applications.xlsx"
always_on_top = False

root = tk.Tk()
root.title("Job Application Logger")
root.geometry("800x600")
root.configure(bg="#f5f5f5")

# Main Frame
main_frame = tk.Frame(root, bg="#f5f5f5")
main_frame.pack(fill=tk.BOTH, expand=True)

# Title
title_label = tk.Label(main_frame, text="Job Application Logger", font=("Arial", 18, "bold"), bg="#f5f5f5")
title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky="n")

# Input Frame
input_frame = tk.Frame(main_frame, bg="#f5f5f5")
input_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

# Input Labels and Entries
tk.Label(input_frame, text="Company:", bg="#f5f5f5").grid(row=0, column=0, sticky="w", padx=10, pady=5)
entry_company = tk.Entry(input_frame)
entry_company.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

tk.Label(input_frame, text="Position:", bg="#f5f5f5").grid(row=1, column=0, sticky="w", padx=10, pady=5)
entry_position = tk.Entry(input_frame)
entry_position.grid(row=1, column=1, padx=10, pady=5, sticky="ew")

tk.Label(input_frame, text="Job Link:", bg="#f5f5f5").grid(row=2, column=0, sticky="w", padx=10, pady=5)
entry_job_link = tk.Entry(input_frame)
entry_job_link.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

tk.Label(input_frame, text="Status:", bg="#f5f5f5").grid(row=3, column=0, sticky="w", padx=10, pady=5)
entry_status = tk.Entry(input_frame)
entry_status.grid(row=3, column=1, padx=10, pady=5, sticky="ew")

tk.Label(input_frame, text="Job Site:", bg="#f5f5f5").grid(row=4, column=0, sticky="w", padx=10, pady=5)
entry_job_site = tk.Entry(input_frame)
entry_job_site.grid(row=4, column=1, padx=10, pady=5, sticky="ew")

submit_button = tk.Button(input_frame, text="Submit", command=submit, bg="#4CAF50", fg="white", font=("Arial", 12))
submit_button.grid(row=5, column=0, columnspan=2, pady=10)

# Treeview Frame
treeview_frame = tk.Frame(main_frame, bg="#f5f5f5")
treeview_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

# Treeview for displaying applications
tree = ttk.Treeview(treeview_frame, columns=("Date Applied", "Company", "Position", "Status"), show="headings")
tree.heading("Date Applied", text="Date Applied")
tree.heading("Company", text="Company")
tree.heading("Position", text="Position")
tree.heading("Status", text="Status")

tree.column("Date Applied", width=120)
tree.column("Company", width=200)
tree.column("Position", width=150)
tree.column("Status", width=100)

tree.pack(fill=tk.BOTH, expand=True)

# Scrollbars
scroll_y = tk.Scrollbar(treeview_frame, orient=tk.VERTICAL, command=tree.yview)
scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
tree.config(yscrollcommand=scroll_y.set)

scroll_x = tk.Scrollbar(treeview_frame, orient=tk.HORIZONTAL, command=tree.xview)
scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
tree.config(xscrollcommand=scroll_x.set)

# Button Frame
button_frame = tk.Frame(main_frame, bg="#f5f5f5")
button_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

# Dropdown menu for selecting field to edit
field_var = tk.StringVar(root)
field_var.set("Select field")
field_menu = tk.OptionMenu(button_frame, field_var, "Company", "Position", "Status", "Date", "Job Link", "Job Site")
field_menu.grid(row=0, column=0, padx=10, pady=5, sticky="ew")

edit_button = tk.Button(button_frame, text="Edit Selected Field", command=edit_entry, bg="#FFC107", fg="black", font=("Arial", 12))
edit_button.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

delete_button = tk.Button(button_frame, text="Delete Selected Entry", command=delete_entry, bg="#F44336", fg="white", font=("Arial", 12))
delete_button.grid(row=0, column=2, padx=10, pady=5, sticky="ew")

open_link_button = tk.Button(button_frame, text="Open Link", command=open_link, bg="#2196F3", fg="white", font=("Arial", 12))
open_link_button.grid(row=0, column=3, padx=10, pady=5, sticky="ew")

toggle_button = tk.Button(button_frame, text="Toggle Always on Top", command=toggle_always_on_top, bg="#607D8B", fg="white", font=("Arial", 12))
toggle_button.grid(row=0, column=4, padx=10, pady=5, sticky="ew")

# Configure grid weights
main_frame.grid_rowconfigure(1, weight=0)  # Input Frame row
main_frame.grid_rowconfigure(2, weight=1)  # Treeview Frame row
main_frame.grid_rowconfigure(3, weight=0)  # Button Frame row
main_frame.grid_columnconfigure(0, weight=1)  # Column for all content

treeview_frame.grid_rowconfigure(0, weight=1)  # Treeview row
treeview_frame.grid_columnconfigure(0, weight=1)  # Treeview column

input_frame.grid_columnconfigure(1, weight=1)  # Input fields expand

update_treeview()  # Initialize treeview with current entries

root.mainloop()
