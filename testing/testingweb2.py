import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime
import webbrowser

# Latest logger

STATUS_COLORS = {
    "Applied": "#FFFFFF",  # White
    "In Progress": "#FFA500",  # Orange
    "Offer": "#008000",  # Green
    "Interview": "#0000FF",  # Blue
    "Rejected": "#FF0000"  # Red
}

def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["ID", "Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"])
    return wb, ws

def submit():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()
    notes = entry_notes.get()
    selected_item = table.selection()
    date_applied = date_entry.get_date().strftime("%d/%m/%Y")

    if company and position and job_link and status and job_site and notes:
        wb, ws = get_workbook(filename)
        if selected_item:  # Update existing record
            item_id = table.item(selected_item)["values"][0]
            for row in ws.iter_rows(min_row=2):
                if row[0].value == item_id:
                    row[1].value = date_applied
                    row[2].value = company
                    row[3].value = position
                    row[4].value = status
                    row[5].value = job_link
                    row[6].value = job_site
                    row[7].value = notes
                    break
        else:  # Add new record
            new_id = ws.max_row  # Using row number as ID
            ws.append([new_id, date_applied, company, position, status, job_link, job_site, notes])
        wb.save(filename)
        messagebox.showinfo("Success", "Application updated successfully!" if selected_item else "Application logged successfully!")
        update_table()
        update_summary()
    else:
        messagebox.showwarning("Input Error", "All fields are required.")

def on_item_select(event):
    selected_item = table.selection()
    if selected_item:
        item = table.item(selected_item)
        values = item['values']
        
        # Update the fields with selected item values
        entry_company.delete(0, tk.END)
        entry_company.insert(0, values[2])
        entry_position.delete(0, tk.END)
        entry_position.insert(0, values[3])
        entry_status.set(values[4])  # Use set() for Combobox
        entry_job_link.delete(0, tk.END)
        entry_job_link.insert(0, values[5])
        entry_job_site.delete(0, tk.END)
        entry_job_site.insert(0, values[6])
        entry_notes.delete(0, tk.END)
        entry_notes.insert(0, values[7])
        
        # Convert date string to datetime.date object and set it in DateEntry
        date_str = values[1]  # Date Applied
        date_obj = datetime.strptime(date_str, "%d/%m/%Y").date()
        date_entry.set_date(date_obj)

# def update_table(filter_text=""):
#     wb, ws = get_workbook(filename)
#     for row in table.get_children():
#         table.delete(row)
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         if not filter_text or any(filter_text.lower() in str(cell).lower() for cell in row):
#             table.insert("", "end", values=row)
#     update_summary()

def update_table(filter_text=""):
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not filter_text or any(filter_text.lower() in str(cell).lower() for cell in row):
            item_id = row[0]
            status = row[4]
            color = STATUS_COLORS.get(status, "#FFFFFF")  # Default to white if status is not found
            table.insert("", "end", iid=item_id, values=row, tags=(status,))
    
    # Apply color tags to rows
    for status, color in STATUS_COLORS.items():
        table.tag_configure(status, background=color)
    
    update_summary()

def search():
    filter_text = search_entry.get()
    update_table(filter_text)

def clear_form():
    entry_company.delete(0, tk.END)
    entry_position.delete(0, tk.END)
    entry_status.set('')  # Reset Combobox selection
    entry_job_link.delete(0, tk.END)
    entry_job_site.delete(0, tk.END)
    entry_notes.delete(0, tk.END)
    date_entry.set_date(datetime.now())

def update_summary():
    wb, ws = get_workbook(filename)
    total = 0
    status_counts = {
        "Applied": 0,
        "In Progress": 0,
        "Offer": 0,
        "Interview": 0,
        "Rejected": 0
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        status = row[4]
        if status in status_counts:
            status_counts[status] += 1
    
    total_label.config(text=f"Total Applications: {total}")
    applied_label.config(text=f"Applied: {status_counts['Applied']}")
    in_progress_label.config(text=f"In Progress: {status_counts['In Progress']}")
    offer_label.config(text=f"Offer: {status_counts['Offer']}")
    interview_label.config(text=f"Interview: {status_counts['Interview']}")
    rejected_label.config(text=f"Rejected: {status_counts['Rejected']}")

def open_link():
    selected_item = table.selection()
    if selected_item:
        item = table.item(selected_item)
        job_link = item['values'][5]  # The Job Link column is index 5
        if job_link:
            webbrowser.open(job_link)
        else:
            messagebox.showwarning("No Link", "No job link available for the selected item.")
    else:
        messagebox.showwarning("No Selection", "Please select a job from the table to open the link.")

filename = "job_app.xlsx"

root = tk.Tk()
root.title("Job Application Tracker")

main_frame = ttk.Frame(root)
main_frame.pack(fill="both", expand=True, padx=10, pady=10)

# Summary Frame
summary_frame = ttk.Frame(main_frame)
summary_frame.grid(row=0, column=0, sticky="nw")

total_label = tk.Label(summary_frame, text="Total Applications: 0", anchor="w")
total_label.pack(fill="x", pady=2)

applied_label = tk.Label(summary_frame, text="Applied: 0", anchor="w")
applied_label.pack(fill="x", pady=2)

in_progress_label = tk.Label(summary_frame, text="In Progress: 0", anchor="w")
in_progress_label.pack(fill="x", pady=2)

offer_label = tk.Label(summary_frame, text="Offer: 0", anchor="w")
offer_label.pack(fill="x", pady=2)

interview_label = tk.Label(summary_frame, text="Interview: 0", anchor="w")
interview_label.pack(fill="x", pady=2)

rejected_label = tk.Label(summary_frame, text="Rejected: 0", anchor="w")
rejected_label.pack(fill="x", pady=2)

# Input Form Frame
input_frame = ttk.Frame(main_frame)
input_frame.grid(row=0, column=1, padx=10, pady=10)

tk.Label(input_frame, text="Date Applied:").grid(row=0, column=0, padx=5, pady=5)
date_entry = DateEntry(input_frame)
date_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Company:").grid(row=1, column=0, padx=5, pady=5)
entry_company = tk.Entry(input_frame)
entry_company.grid(row=1, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Position:").grid(row=2, column=0, padx=5, pady=5)
entry_position = tk.Entry(input_frame)
entry_position.grid(row=2, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Status:").grid(row=3, column=0, padx=5, pady=5)
entry_status = ttk.Combobox(input_frame, values=["Applied", "In Progress", "Offer", "Interview", "Rejected"])
entry_status.grid(row=3, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Job Link:").grid(row=4, column=0, padx=5, pady=5)
entry_job_link = tk.Entry(input_frame)
entry_job_link.grid(row=4, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Job Site:").grid(row=5, column=0, padx=5, pady=5)
entry_job_site = tk.Entry(input_frame)
entry_job_site.grid(row=5, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Notes:").grid(row=6, column=0, padx=5, pady=5)
entry_notes = tk.Entry(input_frame)
entry_notes.grid(row=6, column=1, padx=5, pady=5)

tk.Button(input_frame, text="Submit", command=submit).grid(row=7, column=0, pady=10)

# Add Clear button
tk.Button(input_frame, text="Clear", command=clear_form).grid(row=7, column=1, pady=10)

# Add the Open Link button to the input_frame
tk.Button(input_frame, text="Open Link", command=open_link).grid(row=8, column=0, pady=10, columnspan=2)

# Search Frame
search_frame = ttk.Frame(main_frame)
search_frame.grid(row=1, column=1, padx=10, pady=10)

tk.Label(search_frame, text="Search:").grid(row=0, column=0, padx=5, pady=5)
search_entry = tk.Entry(search_frame)
search_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Button(search_frame, text="Search", command=search).grid(row=0, column=2, padx=5, pady=5)

# Table Frame
table_frame = ttk.Frame(main_frame)
table_frame.grid(row=2, column=0, columnspan=2, pady=10)

table = ttk.Treeview(table_frame, columns=("ID", "Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"), show='headings')
table.pack(fill="both", expand=True)

for col in table["columns"]:
    table.heading(col, text=col)

table.bind("<ButtonRelease-1>", on_item_select)

update_table()
root.mainloop()
