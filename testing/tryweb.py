import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime
import webbrowser

def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"])
    return wb, ws

def update_table():
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        table.insert("", "end", values=row)

def submit():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()
    notes = entry_notes.get()

    if company and position and job_link and status and job_site and notes:
        wb, ws = get_workbook(filename)
        date_applied = date_entry.get_date().strftime("%d/%m/%Y")
        ws.append([date_applied, company, position, status, job_link, job_site, notes])
        wb.save(filename)
        messagebox.showinfo("Success", "Application logged successfully!")
        update_table()
    else:
        messagebox.showwarning("Input Error", "All fields are required.")

filename = "job_applications.xlsx"

root = tk.Tk()
root.title("Job Application Tracker")

add_tab = ttk.Frame(root)
add_tab.pack(fill="both", expand=True)

input_frame = ttk.Frame(add_tab)
input_frame.pack(pady=10, padx=10)

tk.Label(input_frame, text="Date Applied:").grid(row=0, column=0, padx=5, pady=5)
date_entry = DateEntry(input_frame)
date_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Company:").grid(row=1, column=0, padx=5, pady=5)
entry_company = tk.Entry(input_frame)
entry_company.grid(row=1, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Position:").grid(row=2, column=0, padx=5, pady=5)
entry_position = tk.Entry(input_frame)
entry_position.grid(row=2, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Job Link:").grid(row=3, column=0, padx=5, pady=5)
entry_job_link = tk.Entry(input_frame)
entry_job_link.grid(row=3, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Status:").grid(row=4, column=0, padx=5, pady=5)
entry_status = tk.Entry(input_frame)
entry_status.grid(row=4, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Job Site:").grid(row=5, column=0, padx=5, pady=5)
entry_job_site = tk.Entry(input_frame)
entry_job_site.grid(row=5, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Notes:").grid(row=6, column=0, padx=5, pady=5)
entry_notes = tk.Entry(input_frame)
entry_notes.grid(row=6, column=1, padx=5, pady=5)

tk.Button(input_frame, text="Submit", command=submit).grid(row=7, column=0, columnspan=2, pady=10)

table_frame = ttk.Frame(root)
table_frame.pack(fill="both", expand=True)

table = ttk.Treeview(table_frame, columns=("Date Applied", "Company", "Position", "Status", "Job Site", "Notes"), show='headings')
table.pack(fill="both", expand=True)

for col in table["columns"]:
    table.heading(col, text=col)

update_table()
root.mainloop()
