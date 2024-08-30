import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime

def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["ID", "Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"])
    return wb, ws

def submit():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()
    notes = entry_notes.get()
    selected_item = table.selection()
    date_applied = date_entry.get_date().strftime("%d/%m/%Y")

    if company and position and job_link and status and job_site and notes:
        wb, ws = get_workbook(filename)
        if selected_item:  # Update existing record
            item_id = table.item(selected_item)["values"][0]
            for row in ws.iter_rows(min_row=2):
                if row[0].value == item_id:
                    row[1].value = date_applied
                    row[2].value = company
                    row[3].value = position
                    row[4].value = status
                    row[5].value = job_link
                    row[6].value = job_site
                    row[7].value = notes
                    break
        else:  # Add new record
            new_id = ws.max_row  # Using row number as ID
            ws.append([new_id, date_applied, company, position, status, job_link, job_site, notes])
        wb.save(filename)
        messagebox.showinfo("Success", "Application updated successfully!" if selected_item else "Application logged successfully!")
        update_table()
    else:
        messagebox.showwarning("Input Error", "All fields are required.")

def on_item_select(event):
    selected_item = table.selection()
    if selected_item:
        item = table.item(selected_item)
        values = item['values']
        
        # Update the fields with selected item values
        entry_company.delete(0, tk.END)
        entry_company.insert(0, values[2])
        entry_position.delete(0, tk.END)
        entry_position.insert(0, values[3])
        entry_job_link.delete(0, tk.END)
        entry_job_link.insert(0, values[4])
        entry_status.delete(0, tk.END)
        entry_status.insert(0, values[5])
        entry_job_site.delete(0, tk.END)
        entry_job_site.insert(0, values[6])
        entry_notes.delete(0, tk.END)
        entry_notes.insert(0, values[7])
        
        # Convert date string to datetime.date object and set it in DateEntry
        date_str = values[1]  # Date Applied
        date_obj = datetime.strptime(date_str, "%d/%m/%Y").date()
        date_entry.set_date(date_obj)

def update_table(filter_text=""):
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not filter_text or any(filter_text.lower() in str(cell).lower() for cell in row):
            table.insert("", "end", values=row)

def search():
    filter_text = search_entry.get()
    update_table(filter_text)

filename = "testinglist.xlsx"

root = tk.Tk()
root.title("Job Application Tracker")

add_tab = ttk.Frame(root)
add_tab.pack(fill="both", expand=True)

input_frame = ttk.Frame(add_tab)
input_frame.pack(pady=10, padx=10)

tk.Label(input_frame, text="Date Applied:").grid(row=0, column=0, padx=5, pady=5)
date_entry = DateEntry(input_frame)
date_entry.grid(row=0, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Company:").grid(row=1, column=0, padx=5, pady=5)
entry_company = tk.Entry(input_frame)
entry_company.grid(row=1, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Position:").grid(row=2, column=0, padx=5, pady=5)
entry_position = tk.Entry(input_frame)
entry_position.grid(row=2, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Job Link:").grid(row=3, column=0, padx=5, pady=5)
entry_job_link = tk.Entry(input_frame)
entry_job_link.grid(row=3, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Status:").grid(row=4, column=0, padx=5, pady=5)
entry_status = tk.Entry(input_frame)
entry_status.grid(row=4, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Job Site:").grid(row=5, column=0, padx=5, pady=5)
entry_job_site = tk.Entry(input_frame)
entry_job_site.grid(row=5, column=1, padx=5, pady=5)

tk.Label(input_frame, text="Notes:").grid(row=6, column=0, padx=5, pady=5)
entry_notes = tk.Entry(input_frame)
entry_notes.grid(row=6, column=1, padx=5, pady=5)

tk.Button(input_frame, text="Submit", command=submit).grid(row=7, column=0, columnspan=2, pady=10)

search_frame = ttk.Frame(root)
search_frame.pack(pady=10, padx=10)

tk.Label(search_frame, text="Search:").grid(row=0, column=0, padx=5, pady=5)
search_entry = tk.Entry(search_frame)
search_entry.grid(row=0, column=1, padx=5, pady=5)
tk.Button(search_frame, text="Search", command=search).grid(row=0, column=2, padx=5, pady=5)

table_frame = ttk.Frame(root)
table_frame.pack(fill="both", expand=True)

table = ttk.Treeview(table_frame, columns=("ID", "Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"), show='headings')
table.pack(fill="both", expand=True)

for col in table["columns"]:
    table.heading(col, text=col)

table.bind("<ButtonRelease-1>", on_item_select)

update_table()
root.mainloop()
