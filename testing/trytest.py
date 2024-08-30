import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime
import webbrowser

# 7 rows
# Original

# Function to create a new workbook or load an existing one
def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"])
    return wb, ws

# Function to add job application data
def add_job_application(filename, company, position, job_link, status, job_site, notes):
    wb, ws = get_workbook(filename)
    date_applied = date_entry.get_date().strftime("%d/%m/%Y")  # Use DateEntry date picker
    ws.append([date_applied, company, position, status, job_link, job_site, notes])
    wb.save(filename)
    messagebox.showinfo("Success", "Application logged successfully!")

# Function to update the table with entries
def update_table():
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_applied, company, position, status, job_link, job_site, notes = row
        color = status_colors.get(status.lower(), "black")
        table.insert("", "end", values=(date_applied, company, position, status, job_site, notes), tags=(color,))

# Function to handle the form submission
def submit():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()
    notes = entry_notes.get()

    if company and position and job_link and status and job_site and notes:
        add_job_application(filename, company, position, job_link, status, job_site, notes)
        entry_company.delete(0, tk.END)
        entry_position.delete(0, tk.END)
        entry_job_link.delete(0, tk.END)
        entry_status.delete(0, tk.END)
        entry_job_site.delete(0, tk.END)
        entry_notes.delete(0, tk.END)
        update_table()
        update_summary()
    else:
        messagebox.showwarning("Input Error", "All fields are required.")

# Function to edit selected entry
def edit_entry():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to edit.")
        return

    index = table.index(selected_item[0]) + 2  # offset for header row and 0-index
    wb, ws = get_workbook(filename)
    field = field_var.get()

    new_value = None
    if field == "Company":
        new_value = simpledialog.askstring("Edit Company", "Enter new company:")
        if new_value:
            ws.cell(row=index, column=2).value = new_value
    elif field == "Position":
        new_value = simpledialog.askstring("Edit Position", "Enter new position:")
        if new_value:
            ws.cell(row=index, column=3).value = new_value
    elif field == "Job Link":
        new_value = simpledialog.askstring("Edit Job Link", "Enter new job link:")
        if new_value:
            ws.cell(row=index, column=5).value = new_value
    elif field == "Status":
        new_value = simpledialog.askstring("Edit Status", "Enter new status:")
        if new_value:
            ws.cell(row=index, column=4).value = new_value
    elif field == "Job Site":
        new_value = simpledialog.askstring("Edit Job Site", "Enter new job site:")
        if new_value:
            ws.cell(row=index, column=6).value = new_value
    elif field == "Notes":
        new_value = simpledialog.askstring("Edit Notes", "Enter new notes:")
        if new_value:
            ws.cell(row=index, column=7).value = new_value
    elif field == "Date Applied":
        new_value = simpledialog.askstring("Edit Date Applied", "Enter new date (dd/mm/yyyy):")
        if new_value:
            ws.cell(row=index, column=1).value = new_value
    else:
        messagebox.showwarning("Field Error", "Invalid field selected.")
        return

    wb.save(filename)
    messagebox.showinfo("Success", "Entry updated successfully!")
    update_table()

# Function to delete selected entry
def delete_entry():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to delete.")
        return

    confirm = messagebox.askyesno("Delete Confirmation", "Are you sure you want to delete this entry?")
    if confirm:
        index = table.index(selected_item[0]) + 2  # offset for header row and 0-index
        wb, ws = get_workbook(filename)
        ws.delete_rows(index)
        wb.save(filename)
        messagebox.showinfo("Success", "Entry deleted successfully!")
        update_table()
        update_summary()

# Function to open the link of the selected entry
def open_link():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to open.")
        return

    index = table.index(selected_item[0]) + 2  # offset for header row and 0-index
    wb, ws = get_workbook(filename)
    job_link = ws.cell(row=index, column=5).value
    if job_link:
        webbrowser.open(job_link)
    else:
        messagebox.showwarning("Link Error", "No link available for the selected entry.")

# Function to filter the table based on the search query
def filter_table():
    query = search_var.get().lower()
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_applied, company, position, status, job_link, job_site, notes = row
        if query in company.lower() or query in position.lower() or query in status.lower() or query in date_applied.lower():
            color = status_colors.get(status.lower(), "black")
            table.insert("", "end", values=(date_applied, company, position, status, job_site, notes), tags=(color,))

# Function to sort the table by a column
def sort_table(column, reverse):
    data = [(table.set(item, column), item) for item in table.get_children()]
    data.sort(reverse=reverse)
    for index, (val, item) in enumerate(data):
        table.move(item, "", index)
    table.heading(column, command=lambda: sort_table(column, not reverse))

# Function to toggle the stay on top option
def toggle_stay_on_top():
    global stay_on_top
    stay_on_top = not stay_on_top
    root.attributes('-topmost', stay_on_top)
    stay_on_top_button.config(bg="green" if stay_on_top else "red")

# Function to update the summary section
def update_summary():
    wb, ws = get_workbook(filename)
    total_applications = ws.max_row - 1  # exclude header row
    statuses = [row[3] for row in ws.iter_rows(min_row=2, values_only=True)]
    interviews = statuses.count("Interview")
    offers = statuses.count("Offer")
    rejections = statuses.count("Rejected")
    summary_var.set(f"Total Applications: {total_applications}\nInterviews: {interviews}\nOffers: {offers}\nRejections: {rejections}")

# Function to resize window based on tab content
def resize_window(event):
    if notebook.index(notebook.select()) == 0:  # Add Application tab
        root.update_idletasks()
        root.geometry(f"{input_frame.winfo_reqwidth() + 20}x{input_frame.winfo_reqheight() + 100}")
    elif notebook.index(notebook.select()) == 1:  # Application List tab
        root.update_idletasks()
        root.geometry(f"{table_frame.winfo_reqwidth() + 20}x{table_frame.winfo_reqheight() + 100}")

# GUI setup
filename = "job_applications.xlsx"
root = tk.Tk()
root.title("Job Application Logger")
root.geometry("900x600")
root.configure(bg="#f5f5f5")

notebook = ttk.Notebook(root)
input_tab = ttk.Frame(notebook)
table_tab = ttk.Frame(notebook)
notebook.add(input_tab, text="Add Application")
notebook.add(table_tab, text="Application List")
notebook.pack(expand=True, fill="both")

# Frame for input form
input_frame = tk.Frame(input_tab, padx=10, pady=10, bg="#f5f5f5")
input_frame.pack(fill="both", expand=True)

tk.Label(input_frame, text="Company:", bg="#f5f5f5").grid(row=0, column=0, sticky="w")
entry_company = tk.Entry(input_frame, width=40)
entry_company.grid(row=0, column=1, sticky="ew")

tk.Label(input_frame, text="Position:", bg="#f5f5f5").grid(row=1, column=0, sticky="w")
entry_position = tk.Entry(input_frame, width=40)
entry_position.grid(row=1, column=1, sticky="ew")

tk.Label(input_frame, text="Job Link:", bg="#f5f5f5").grid(row=2, column=0, sticky="w")
entry_job_link = tk.Entry(input_frame, width=40)
entry_job_link.grid(row=2, column=1, sticky="ew")

tk.Label(input_frame, text="Status:", bg="#f5f5f5").grid(row=3, column=0, sticky="w")
entry_status = tk.Entry(input_frame, width=40)
entry_status.grid(row=3, column=1, sticky="ew")

tk.Label(input_frame, text="Job Site:", bg="#f5f5f5").grid(row=4, column=0, sticky="w")
entry_job_site = tk.Entry(input_frame, width=40)
entry_job_site.grid(row=4, column=1, sticky="ew")

tk.Label(input_frame, text="Notes:", bg="#f5f5f5").grid(row=5, column=0, sticky="w")
entry_notes = tk.Entry(input_frame, width=40)
entry_notes.grid(row=5, column=1, sticky="ew")

tk.Label(input_frame, text="Date Applied:", bg="#f5f5f5").grid(row=6, column=0, sticky="w")
date_entry = DateEntry(input_frame, width=37, background="darkblue", foreground="white", date_pattern="dd/mm/yyyy")
date_entry.grid(row=6, column=1, sticky="ew")

button_frame = tk.Frame(input_tab, pady=10, bg="#f5f5f5")
button_frame.pack(fill="x")

tk.Button(button_frame, text="Add Application", command=submit).pack(side="left", padx=5)
stay_on_top_button = tk.Button(button_frame, text="Stay On Top", bg="red", command=toggle_stay_on_top)
stay_on_top_button.pack(side="left", padx=5)

search_frame = tk.Frame(table_tab, padx=10, pady=10, bg="#f5f5f5")
search_frame.grid(row=0, column=0, sticky="ew")

tk.Label(search_frame, text="Search:", bg="#f5f5f5").grid(row=0, column=0, sticky="w")
search_var = tk.StringVar()
search_entry = tk.Entry(search_frame, textvariable=search_var, width=40)
search_entry.grid(row=0, column=1, padx=5)
search_button = tk.Button(search_frame, text="Search", command=filter_table)
search_button.grid(row=0, column=2, padx=5)

summary_frame = tk.Frame(table_tab, padx=10, pady=10, bg="#f5f5f5")
summary_frame.grid(row=1, column=0, sticky="ew")

summary_var = tk.StringVar()
summary_label = tk.Label(summary_frame, textvariable=summary_var, justify="left", bg="#f5f5f5")
summary_label.grid(row=0, column=0, sticky="w")

table_frame = tk.Frame(table_tab, padx=10, pady=10, bg="#f5f5f5")
table_frame.grid(row=2, column=0, sticky="nsew")

columns = ("date_applied", "company", "position", "status", "job_site", "notes")
table = ttk.Treeview(table_frame, columns=columns, show="headings")
table.heading("date_applied", text="Date Applied", command=lambda: sort_table("date_applied", False))
table.heading("company", text="Company", command=lambda: sort_table("company", False))
table.heading("position", text="Position", command=lambda: sort_table("position", False))
table.heading("status", text="Status", command=lambda: sort_table("status", False))
table.heading("job_site", text="Job Site")
table.heading("notes", text="Notes")

status_colors = {
    "interview": "blue",
    "offer": "green",
    "rejected": "red",
    "in progress": "orange"  # Add this line
}
table.tag_configure("blue", foreground="blue")
table.tag_configure("green", foreground="green")
table.tag_configure("red", foreground="red")
table.tag_configure("orange", foreground="orange")  # Add this line

table.pack(fill="both", expand=True)

edit_frame = tk.Frame(table_tab, pady=10, bg="#f5f5f5")
edit_frame.grid(row=3, column=0, sticky="ew")

field_var = tk.StringVar(value="Company")
fields = ["Company", "Position", "Job Link", "Status", "Job Site", "Notes", "Date Applied"]
field_menu = tk.OptionMenu(edit_frame, field_var, *fields)
field_menu.grid(row=0, column=0, padx=5)

edit_button = tk.Button(edit_frame, text="Edit Selected", command=edit_entry)
edit_button.grid(row=0, column=1, padx=5)
delete_button = tk.Button(edit_frame, text="Delete Selected", command=delete_entry)
delete_button.grid(row=0, column=2, padx=5)
open_link_button = tk.Button(edit_frame, text="Open Link", command=open_link)
open_link_button.grid(row=0, column=3, padx=5)

stay_on_top = False

# Configure column and row weights for resizing
input_tab.grid_rowconfigure(0, weight=1)
input_tab.grid_columnconfigure(0, weight=1)
table_tab.grid_rowconfigure(2, weight=1)
table_tab.grid_columnconfigure(0, weight=1)

# Initial table and summary update
update_table()
update_summary()

# Bind the notebook tab change event to resize the window
notebook.bind("<<NotebookTabChanged>>", resize_window)

# Initial resize based on the default tab
resize_window(None)

root.mainloop()
