import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
import webbrowser

# Function to create a new workbook or load an existing one
def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"])
    return wb, ws

# Function to add job application data
def add_job_application(filename, company, position, job_link, status, job_site, notes):
    wb, ws = get_workbook(filename)
    date_applied = date_entry.get_date().strftime("%d/%m/%Y")  # Use DateEntry date picker
    ws.append([date_applied, company, position, status, job_link, job_site, notes])
    wb.save(filename)
    messagebox.showinfo("Success", "Application logged successfully!")

# Function to update the table with entries
def update_table():
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_applied, company, position, status, job_link, job_site, notes = row
        color = status_colors.get(status.lower(), "black")
        table.insert("", "end", values=(date_applied, company, position, status, job_site, notes), tags=(color,))

# Function to handle the form submission
def submit():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()
    notes = entry_notes.get()

    if company and position and job_link and status and job_site and notes:
        add_job_application(filename, company, position, job_link, status, job_site, notes)
        entry_company.delete(0, tk.END)
        entry_position.delete(0, tk.END)
        entry_job_link.delete(0, tk.END)
        entry_status.delete(0, tk.END)
        entry_job_site.delete(0, tk.END)
        entry_notes.delete(0, tk.END)
        update_table()
        update_summary()
    else:
        messagebox.showwarning("Input Error", "All fields are required.")

def edit_entry():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to edit.")
        return

    selected_values = table.item(selected_item[0])['values']  # Get the values of the selected row
    date_applied, company, position, status, job_site, notes = selected_values

    wb, ws = get_workbook(filename)

    # Find the exact row in the Excel sheet that matches the selected item
    for row in ws.iter_rows(min_row=2, values_only=False):
        if (row[0].value == date_applied and
            row[1].value == company and
            row[2].value == position and
            row[3].value == status and
            row[4].value == job_site and
            row[5].value == notes):
            row_to_edit = row[0].row
            break
    else:
        messagebox.showerror("Edit Error", "The selected entry was not found in the workbook.")
        return

    field = field_var.get()
    new_value = None

    if field == "Company":
        new_value = simpledialog.askstring("Edit Company", "Enter new company:")
        if new_value:
            ws.cell(row=row_to_edit, column=2).value = new_value
    elif field == "Position":
        new_value = simpledialog.askstring("Edit Position", "Enter new position:")
        if new_value:
            ws.cell(row=row_to_edit, column=3).value = new_value
    elif field == "Job Link":
        new_value = simpledialog.askstring("Edit Job Link", "Enter new job link:")
        if new_value:
            ws.cell(row=row_to_edit, column=5).value = new_value
    elif field == "Status":
        new_value = simpledialog.askstring("Edit Status", "Enter new status:")
        if new_value:
            ws.cell(row=row_to_edit, column=4).value = new_value
    elif field == "Job Site":
        new_value = simpledialog.askstring("Edit Job Site", "Enter new job site:")
        if new_value:
            ws.cell(row=row_to_edit, column=6).value = new_value
    elif field == "Notes":
        new_value = simpledialog.askstring("Edit Notes", "Enter new notes:")
        if new_value:
            ws.cell(row=row_to_edit, column=7).value = new_value
    elif field == "Date Applied":
        new_value = simpledialog.askstring("Edit Date Applied", "Enter new date (dd/mm/yyyy):")
        if new_value:
            ws.cell(row=row_to_edit, column=1).value = new_value
    else:
        messagebox.showwarning("Field Error", "Invalid field selected.")
        return

    wb.save(filename)
    messagebox.showinfo("Success", "Entry updated successfully!")
    update_table()

# Function to delete selected entry
def delete_entry():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to delete.")
        return

    confirm = messagebox.askyesno("Delete Confirmation", "Are you sure you want to delete this entry?")
    if confirm:
        selected_values = table.item(selected_item[0])['values']
        date_applied, company, position, status, job_site, notes = selected_values

        wb, ws = get_workbook(filename)
        for row in ws.iter_rows(min_row=2, values_only=False):
            if (
                row[0].value == date_applied and
                row[1].value == company and
                row[2].value == position and
                row[3].value == status and
                row[4].value == job_site and
                row[5].value == notes
            ):
                ws.delete_rows(row[0].row)
                break

        wb.save(filename)
        messagebox.showinfo("Success", "Entry deleted successfully!")
        update_table()
        update_summary()

# Function to open the link of the selected entry
def open_link():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to open.")
        return

    selected_values = table.item(selected_item[0])['values']
    date_applied, company, position, status, job_site, notes = selected_values

    wb, ws = get_workbook(filename)
    for row in ws.iter_rows(min_row=2, values_only=False):
        if (
            row[0].value == date_applied and
            row[1].value == company and
            row[2].value == position and
            row[3].value == status and
            row[4].value == job_site and
            row[5].value == notes
        ):
            job_link = row[4].value
            break
    else:
        job_link = None

    if job_link:
        webbrowser.open(job_link)
    else:
        messagebox.showwarning("Link Error", "No link available for the selected entry.")

# Function to filter the table based on the search query
def filter_table():
    query = search_var.get().lower()
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_applied, company, position, status, job_link, job_site, notes = row
        if query in company.lower() or query in position.lower() or query in status.lower() or query in date_applied.lower():
            color = status_colors.get(status.lower(), "black")
            table.insert("", "end", values=(date_applied, company, position, status, job_site, notes), tags=(color,))

# Function to sort the table by a column
def sort_table(column, reverse):
    data = [(table.set(item, column), item) for item in table.get_children()]
    data.sort(reverse=reverse)
    for index, (val, item) in enumerate(data):
        table.move(item, "", index)
    table.heading(column, command=lambda: sort_table(column, not reverse))

# Function to toggle the stay on top option
def toggle_stay_on_top():
    global stay_on_top
    stay_on_top = not stay_on_top
    root.attributes('-topmost', stay_on_top)
    stay_on_top_button.config(bg="green" if stay_on_top else "red")

# Function to update the summary section
def update_summary():
    wb, ws = get_workbook(filename)
    total_applications = ws.max_row - 1  # exclude header row
    statuses = [row[3] for row in ws.iter_rows(min_row=2, values_only=True)]
    interviews = statuses.count("Interview")
    offers = statuses.count("Offer")
    rejections = statuses.count("Rejected")
    summary_var.set(f"Total Applications: {total_applications}\nInterviews: {interviews}\nOffers: {offers}\nRejections: {rejections}")

# Function to resize window based on the selected tab
def resize_window(event):
    selected_tab = event.widget.tab(event.widget.index("current"))["text"]
    if selected_tab == "Summary":
        root.geometry("400x400")
    else:
        root.geometry("800x600")

# Set up the main application window
root = tk.Tk()
root.title("Job Application Tracker")
root.geometry("800x600")
root.attributes('-topmost', True)  # Set to stay on top by default
stay_on_top = True

filename = "jobtest.xlsx"

status_colors = {
    "applied": "black",
    "interview": "blue",
    "offer": "green",
    "rejected": "red",
    "in progress": "orange"
}

# Create tabs
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill="both")

tab1 = ttk.Frame(notebook)
tab2 = ttk.Frame(notebook)
tab3 = ttk.Frame(notebook)
notebook.add(tab1, text="Add Application")
notebook.add(tab2, text="Application List")
notebook.add(tab3, text="Summary")

notebook.bind("<<NotebookTabChanged>>", resize_window)

# Add Application tab
label_company = tk.Label(tab1, text="Company:")
label_company.grid(row=0, column=0, padx=10, pady=10)
entry_company = tk.Entry(tab1)
entry_company.grid(row=0, column=1, padx=10, pady=10)

label_position = tk.Label(tab1, text="Position:")
label_position.grid(row=1, column=0, padx=10, pady=10)
entry_position = tk.Entry(tab1)
entry_position.grid(row=1, column=1, padx=10, pady=10)

label_job_link = tk.Label(tab1, text="Job Link:")
label_job_link.grid(row=2, column=0, padx=10, pady=10)
entry_job_link = tk.Entry(tab1)
entry_job_link.grid(row=2, column=1, padx=10, pady=10)

label_status = tk.Label(tab1, text="Status:")
label_status.grid(row=3, column=0, padx=10, pady=10)
entry_status = ttk.Combobox(tab1, values=["Applied", "Interview", "Offer", "Rejected", "In Progress"])
entry_status.grid(row=3, column=1, padx=10, pady=10)

label_job_site = tk.Label(tab1, text="Job Site:")
label_job_site.grid(row=4, column=0, padx=10, pady=10)
entry_job_site = tk.Entry(tab1)
entry_job_site.grid(row=4, column=1, padx=10, pady=10)

label_notes = tk.Label(tab1, text="Notes:")
label_notes.grid(row=5, column=0, padx=10, pady=10)
entry_notes = tk.Entry(tab1)
entry_notes.grid(row=5, column=1, padx=10, pady=10)

label_date_applied = tk.Label(tab1, text="Date Applied:")
label_date_applied.grid(row=6, column=0, padx=10, pady=10)
date_entry = DateEntry(tab1, date_pattern="dd/mm/yyyy")
date_entry.grid(row=6, column=1, padx=10, pady=10)

submit_button = tk.Button(tab1, text="Submit", command=submit)
submit_button.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

# Application List tab
table = ttk.Treeview(tab2, columns=("Date Applied", "Company", "Position", "Status", "Job Site", "Notes"), show="headings")
table.heading("Date Applied", text="Date Applied", command=lambda: sort_table("Date Applied", False))
table.heading("Company", text="Company", command=lambda: sort_table("Company", False))
table.heading("Position", text="Position", command=lambda: sort_table("Position", False))
table.heading("Status", text="Status", command=lambda: sort_table("Status", False))
table.heading("Job Site", text="Job Site", command=lambda: sort_table("Job Site", False))
table.heading("Notes", text="Notes", command=lambda: sort_table("Notes", False))
table.tag_configure("blue", foreground="blue")
table.tag_configure("green", foreground="green")
table.tag_configure("red", foreground="red")
table.tag_configure("orange", foreground="orange")
table.pack(expand=True, fill="both")

scrollbar = ttk.Scrollbar(tab2, orient="vertical", command=table.yview)
table.configure(yscroll=scrollbar.set)
scrollbar.pack(side="right", fill="y")

# Buttons for editing and deleting
edit_button = tk.Button(tab2, text="Edit", command=edit_entry)
edit_button.pack(side="left", padx=10, pady=10)

delete_button = tk.Button(tab2, text="Delete", command=delete_entry)
delete_button.pack(side="left", padx=10, pady=10)

link_button = tk.Button(tab2, text="Open Link", command=open_link)
link_button.pack(side="left", padx=10, pady=10)

search_var = tk.StringVar()
search_entry = tk.Entry(tab2, textvariable=search_var)
search_entry.pack(side="left", padx=10, pady=10)

search_button = tk.Button(tab2, text="Search", command=filter_table)
search_button.pack(side="left", padx=10, pady=10)

field_var = tk.StringVar(value="Company")
field_menu = ttk.Combobox(tab2, textvariable=field_var, values=["Company", "Position", "Job Link", "Status", "Job Site", "Notes", "Date Applied"])
field_menu.pack(side="left", padx=10, pady=10)

stay_on_top_button = tk.Button(tab1, text="Stay on Top", command=toggle_stay_on_top, bg="green")
stay_on_top_button.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

# Summary tab
summary_label = tk.Label(tab3, text="Summary")
summary_label.pack()

summary_var = tk.StringVar()
summary_text = tk.Label(tab3, textvariable=summary_var, justify="left")
summary_text.pack()

# Initialize the table and summary
update_table()
update_summary()

root.mainloop()
