import tkinter as tk
from tkinter import messagebox, ttk, simpledialog
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime
import webbrowser

# Constants for colors
STATUS_COLORS = {
    "Applied": "#FFFFFF",  # White
    "In Progress": "#FFA500",  # Orange
    "Offer": "#008000",  # Green
    "Interview": "#0000FF",  # Blue
    "Rejected": "#FF0000"  # Red
}

def get_workbook(filename):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        ws.append(["ID", "Date Applied", "Company", "Position", "Status", "Job Link", "Job Site", "Notes"])
    return wb, ws

def add_job_application():
    company = entry_company.get()
    position = entry_position.get()
    job_link = entry_job_link.get()
    status = entry_status.get()
    job_site = entry_job_site.get()
    notes = entry_notes.get()
    date_applied = date_entry.get_date().strftime("%d/%m/%Y")

    if company and position and job_link and status and job_site and notes:
        wb, ws = get_workbook(filename)
        new_id = ws.max_row  # Using row number as ID
        ws.append([new_id, date_applied, company, position, status, job_link, job_site, notes])
        wb.save(filename)
        messagebox.showinfo("Success", "Application logged successfully!")
        entry_company.delete(0, tk.END)
        entry_position.delete(0, tk.END)
        entry_job_link.delete(0, tk.END)
        entry_status.delete(0, tk.END)
        entry_job_site.delete(0, tk.END)
        entry_notes.delete(0, tk.END)
        update_table()
        update_summary()
    else:
        messagebox.showerror("Error", "Please fill in all fields!")

def update_table(filter_text=""):
    wb, ws = get_workbook(filename)
    for row in table.get_children():
        table.delete(row)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not filter_text or any(filter_text.lower() in str(cell).lower() for cell in row):
            item_id = row[0]
            status = row[4]
            color = STATUS_COLORS.get(status, "#FFFFFF")  # Default to white if status is not found
            iidd,date_applied, company, position, status, job_link, job_site, notes = row
            table.insert("", "end", iid=item_id, values=(date_applied, company, position, status, job_site, notes), tags=(status,))
    
    # Apply color tags to rows
    for status, color in STATUS_COLORS.items():
        table.tag_configure(status, background=color)
    
    update_summary()

def update_summary():
    wb, ws = get_workbook(filename)
    total = 0
    status_counts = {
        "Applied": 0,
        "In Progress": 0,
        "Offer": 0,
        "Interview": 0,
        "Rejected": 0
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        status = row[4]
        if status in status_counts:
            status_counts[status] += 1
    
    summary_var.set(f"Total Applications: {total}\nApplied: {status_counts['Applied']}\nIn Progress: {status_counts['In Progress']}\nInterviews: {status_counts['Interview']}\nOffers: {status_counts['Offer']}\nRejections: {status_counts['Rejected']}")

def toggle_stay_on_top():
    global stay_on_top
    stay_on_top = not stay_on_top
    root.attributes('-topmost', stay_on_top)
    stay_on_top_button.config(bg="green" if stay_on_top else "red")

def search():
    filter_text = search_entry.get()
    update_table(filter_text)

def sort_table(column, reverse):
    data = [(table.set(item, column), item) for item in table.get_children()]
    data.sort(reverse=reverse)
    for index, (val, item) in enumerate(data):
        table.move(item, "", index)
    table.heading(column, command=lambda: sort_table(column, not reverse))

def open_link():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to open.")
        return

    # Get the ID of the selected item
    item_id = int(selected_item[0])

    # Open the workbook and worksheet
    wb, ws = get_workbook(filename)

    # Iterate through the rows to find the matching ID
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_id = row[0].value  # Assuming the ID is in the first column (A)
        if row_id == item_id:
            job_link = row[5].value  # Assuming job link is in the 6th column (F)
            if job_link:
                webbrowser.open(job_link)
            else:
                messagebox.showwarning("Link Error", "No link available for the selected entry.")
            break
    else:
        messagebox.showwarning("ID Error", "The selected item ID is invalid.")

def edit_entry():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to edit.")
        return

    # Get the ID of the selected item
    item_id = int(selected_item[0])
    
    # Open the workbook and worksheet
    wb, ws = get_workbook(filename)

    # Find the row with the matching ID
    for row in ws.iter_rows(min_row=2):
        row_id = row[0].value  # Assuming the ID is in the first column (A)
        if row_id == item_id:
            field = field_var.get()

            new_value = None
            if field == "Company":
                new_value = simpledialog.askstring("Edit Company", "Enter new company:")
                if new_value:
                    ws.cell(row=row[0].row, column=3).value = new_value  # Company is in the 3rd column
            elif field == "Position":
                new_value = simpledialog.askstring("Edit Position", "Enter new position:")
                if new_value:
                    ws.cell(row=row[0].row, column=4).value = new_value  # Position is in the 4th column
            elif field == "Job Link":
                new_value = simpledialog.askstring("Edit Job Link", "Enter new job link:")
                if new_value:
                    ws.cell(row=row[0].row, column=6).value = new_value  # Job Link is in the 6th column
            elif field == "Status":
                new_value = simpledialog.askstring("Edit Status", "Enter new status:")
                if new_value:
                    ws.cell(row=row[0].row, column=5).value = new_value  # Status is in the 5th column
            elif field == "Job Site":
                new_value = simpledialog.askstring("Edit Job Site", "Enter new job site:")
                if new_value:
                    ws.cell(row=row[0].row, column=7).value = new_value  # Job Site is in the 7th column
            elif field == "Notes":
                new_value = simpledialog.askstring("Edit Notes", "Enter new notes:")
                if new_value:
                    ws.cell(row=row[0].row, column=8).value = new_value  # Notes is in the 8th column
            elif field == "Date Applied":
                new_value = simpledialog.askstring("Edit Date Applied", "Enter new date (dd/mm/yyyy):")
                if new_value:
                    try:
                        datetime.strptime(new_value, "%d/%m/%Y")  # Validate date format
                        ws.cell(row=row[0].row, column=2).value = new_value  # Date Applied is in the 2nd column
                    except ValueError:
                        messagebox.showerror("Date Error", "Invalid date format. Please enter as dd/mm/yyyy.")
                        return
            else:
                messagebox.showwarning("Field Error", "Invalid field selected.")
                return

            wb.save(filename)
            messagebox.showinfo("Success", "Entry updated successfully!")
            update_table()
            update_summary()
            break
    else:
        messagebox.showwarning("ID Error", "The selected item ID is invalid.")

def delete_entry():
    selected_item = table.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select an entry to delete.")
        return

    # Get the ID of the selected item
    item_id = int(selected_item[0])

    # Confirm deletion
    confirm = messagebox.askyesno("Delete Confirmation", "Are you sure you want to delete this entry?")
    if not confirm:
        return
    
    # Open the workbook and worksheet
    wb, ws = get_workbook(filename)

    # Find the row with the matching ID and delete it
    for row in ws.iter_rows(min_row=2):
        row_id = row[0].value  # Assuming the ID is in the first column (A)
        if row_id == item_id:
            ws.delete_rows(row[0].row)  # Delete the row by its index
            wb.save(filename)
            messagebox.showinfo("Success", "Entry deleted successfully!")
            update_table()
            update_summary()
            break
    else:
        messagebox.showwarning("ID Error", "The selected item ID is invalid.")


def resize_window(event):
    if notebook.index(notebook.select()) == 0:  # Add Application tab
        root.update_idletasks()
        root.geometry(f"{input_frame.winfo_reqwidth() + 20}x{input_frame.winfo_reqheight() + 100}")
    elif notebook.index(notebook.select()) == 1:  # Application List tab
        root.update_idletasks()
        root.geometry(f"{table_frame.winfo_reqwidth() + 20}x{table_frame.winfo_reqheight() + 100}")

#----------------------------------------------------------------------------------------------------------------------
# GUI setup
filename = "job_app.xlsx"
root = tk.Tk()
root.title("Job Application Logger")
root.geometry("900x600")
root.configure(bg="#f5f5f5")

notebook = ttk.Notebook(root)
input_tab = ttk.Frame(notebook)
table_tab = ttk.Frame(notebook)
notebook.add(input_tab, text="Add Application")
notebook.add(table_tab, text="Application List")
notebook.pack(expand=True, fill="both")

# Frame for input form
input_frame = tk.Frame(input_tab, padx=10, pady=10, bg="#f5f5f5")
input_frame.pack(fill="both", expand=True)

tk.Label(input_frame, text="Company:", bg="#f5f5f5").grid(row=0, column=0, sticky="w")
entry_company = tk.Entry(input_frame, width=40)
entry_company.grid(row=0, column=1, sticky="ew")

tk.Label(input_frame, text="Position:", bg="#f5f5f5").grid(row=1, column=0, sticky="w")
entry_position = tk.Entry(input_frame, width=40)
entry_position.grid(row=1, column=1, sticky="ew")

tk.Label(input_frame, text="Job Link:", bg="#f5f5f5").grid(row=2, column=0, sticky="w")
entry_job_link = tk.Entry(input_frame, width=40)
entry_job_link.grid(row=2, column=1, sticky="ew")

tk.Label(input_frame, text="Status:", bg="#f5f5f5").grid(row=3, column=0, sticky="w")
entry_status = ttk.Combobox(input_frame, values=["Applied", "In Progress", "Offer", "Interview", "Rejected"])
entry_status.grid(row=3, column=1, sticky="ew")

tk.Label(input_frame, text="Job Site:", bg="#f5f5f5").grid(row=4, column=0, sticky="w")
entry_job_site = tk.Entry(input_frame, width=40)
entry_job_site.grid(row=4, column=1, sticky="ew")

tk.Label(input_frame, text="Notes:", bg="#f5f5f5").grid(row=5, column=0, sticky="w")
entry_notes = tk.Entry(input_frame, width=40)
entry_notes.grid(row=5, column=1, sticky="ew")

tk.Label(input_frame, text="Date Applied:", bg="#f5f5f5").grid(row=6, column=0, sticky="w")
date_entry = DateEntry(input_frame, width=37, background="darkblue", foreground="white", date_pattern="dd/mm/yyyy")
date_entry.grid(row=6, column=1, sticky="ew")

button_frame = tk.Frame(input_tab, pady=10, bg="#f5f5f5")
button_frame.pack(fill="x")

tk.Button(button_frame, text="Add Application", command=add_job_application).pack(side="left", padx=5)
stay_on_top_button = tk.Button(button_frame, text="Stay On Top", bg="red", command=toggle_stay_on_top)
stay_on_top_button.pack(side="left", padx=5)

search_frame = tk.Frame(table_tab, padx=10, pady=10, bg="#f5f5f5")
search_frame.grid(row=0, column=0, sticky="ew")

tk.Label(search_frame, text="Search:", bg="#f5f5f5").grid(row=0, column=0, sticky="w")
search_var = tk.StringVar()
search_entry = tk.Entry(search_frame, textvariable=search_var, width=40)
search_entry.grid(row=0, column=1, padx=5)
search_button = tk.Button(search_frame, text="Search", command=search)
search_button.grid(row=0, column=2, padx=5)

summary_frame = tk.Frame(table_tab, padx=10, pady=10, bg="#f5f5f5")
summary_frame.grid(row=1, column=0, sticky="ew")

summary_var = tk.StringVar()
summary_label = tk.Label(summary_frame, textvariable=summary_var, justify="left", bg="#f5f5f5")
summary_label.grid(row=0, column=0, sticky="w")

table_frame = tk.Frame(table_tab, padx=10, pady=10, bg="#f5f5f5")
table_frame.grid(row=2, column=0, sticky="nsew")

columns = ("date_applied", "company", "position", "status", "job_site", "notes")
table = ttk.Treeview(table_frame, columns=columns, show="headings")
table.heading("date_applied", text="Date Applied", command=lambda: sort_table("date_applied", False))
table.heading("company", text="Company", command=lambda: sort_table("company", False))
table.heading("position", text="Position", command=lambda: sort_table("position", False))
table.heading("status", text="Status", command=lambda: sort_table("status", False))
table.heading("job_site", text="Job Site")
table.heading("notes", text="Notes")

status_colors = {
    "interview": "blue",
    "offer": "green",
    "rejected": "red",
    "in progress": "orange"  # Add this line
}
table.tag_configure("blue", foreground="blue")
table.tag_configure("green", foreground="green")
table.tag_configure("red", foreground="red")
table.tag_configure("orange", foreground="orange")  # Add this line

table.pack(fill="both", expand=True)

edit_frame = tk.Frame(table_tab, pady=10, bg="#f5f5f5")
edit_frame.grid(row=3, column=0, sticky="ew")

field_var = tk.StringVar(value="Company")
fields = ["Company", "Position", "Job Link", "Status", "Job Site", "Notes", "Date Applied"]
field_menu = tk.OptionMenu(edit_frame, field_var, *fields)
field_menu.grid(row=0, column=0, padx=5)

edit_button = tk.Button(edit_frame, text="Edit Selected", command=edit_entry)
edit_button.grid(row=0, column=1, padx=5)
delete_button = tk.Button(edit_frame, text="Delete Selected", command=delete_entry)
delete_button.grid(row=0, column=2, padx=5)
open_link_button = tk.Button(edit_frame, text="Open Link", command=open_link)
open_link_button.grid(row=0, column=3, padx=5)

stay_on_top = False

# Configure column and row weights for resizing
input_tab.grid_rowconfigure(0, weight=1)
input_tab.grid_columnconfigure(0, weight=1)
table_tab.grid_rowconfigure(2, weight=1)
table_tab.grid_columnconfigure(0, weight=1)

# Initial table and summary update
update_table()
update_summary()

# Bind the notebook tab change event to resize the window
notebook.bind("<<NotebookTabChanged>>", resize_window)

# Initial resize based on the default tab
resize_window(None)

root.mainloop()